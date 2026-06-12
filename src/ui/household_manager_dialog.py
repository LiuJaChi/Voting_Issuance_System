"""
住戶管理對話框 - 支持 .xlsx/.csv 格式
欄位支持：戶號 | 戶名 | 面積（坪）
"""
import csv
from pathlib import Path
from typing import List, Dict

from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit,
    QPushButton, QTableWidget, QTableWidgetItem, QMessageBox,
    QHeaderView, QFileDialog
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QFont

from src.backend.database import Database

try:
    import openpyxl
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False


class HouseholdManagerDialog(QDialog):
    """住戶管理對話框 - 支持住戶導入和報到資料導出"""
    
    def __init__(self, parent=None):
        """初始化住戶管理對話框"""
        super().__init__(parent)
        self.setWindowTitle("住戶管理")
        self.setGeometry(100, 100, 1000, 600)
        
        self.db = Database()
        
        self.init_ui()
        self.load_households()
    
    def init_ui(self):
        """初始化用戶界面"""
        main_layout = QVBoxLayout()
        
        # 標題
        title = QLabel("住戶管理")
        title_font = QFont()
        title_font.setPointSize(14)
        title_font.setBold(True)
        title.setFont(title_font)
        main_layout.addWidget(title)
        
        # 搜索和過濾區域
        search_layout = QHBoxLayout()
        search_layout.addWidget(QLabel("搜尋戶號:"))
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("輸入戶號進行搜尋...")
        self.search_input.textChanged.connect(self.filter_households)
        search_layout.addWidget(self.search_input)
        search_layout.addStretch()
        main_layout.addLayout(search_layout)
        
        # 住戶表格 - 欄位順序：戶號 | 戶名 | 面積（坪）
        self.household_table = QTableWidget()
        self.household_table.setColumnCount(3)
        self.household_table.setHorizontalHeaderLabels(["戶號", "戶名", "面積（坪）"])
        
        for col in range(3):
            self.household_table.horizontalHeader().setSectionResizeMode(
                col, QHeaderView.ResizeMode.Stretch
            )
        
        self.household_table.setSelectionBehavior(
            QTableWidget.SelectionBehavior.SelectRows
        )
        self.household_table.setSelectionMode(
            QTableWidget.SelectionMode.SingleSelection
        )
        main_layout.addWidget(self.household_table)
        
        # 按鈕區域
        button_layout = QHBoxLayout()
        
        # 導入按鈕
        import_button = QPushButton("導入住戶（.xlsx/.csv）")
        import_button.clicked.connect(self.import_households)
        button_layout.addWidget(import_button)
        
        # 導出報到資料按鈕
        export_button = QPushButton("導出報到資料（.xlsx）")
        export_button.clicked.connect(self.export_check_in_data)
        button_layout.addWidget(export_button)
        
        button_layout.addSpacing(20)
        
        # 刷新按鈕
        refresh_button = QPushButton("刷新")
        refresh_button.clicked.connect(self.load_households)
        button_layout.addWidget(refresh_button)
        
        button_layout.addStretch()
        
        # 關閉按鈕
        close_button = QPushButton("關閉")
        close_button.clicked.connect(self.accept)
        button_layout.addWidget(close_button)
        
        main_layout.addLayout(button_layout)
        
        self.setLayout(main_layout)
        
        # 保存所有住戶數據以供搜尋
        self.all_households = []
    
    def load_households(self):
        """加載所有住戶"""
        self.all_households = []
        
        households = self.db.get_all_households()
        for household in households:
            self.all_households.append({
                'household_id': household['household_id'],
                'name': household['name'],
                'share_amount': household.get('share_amount', 0.0),
            })
        
        self.refresh_table(self.all_households)
    
    def refresh_table(self, households: List[Dict]):
        """刷新表格顯示 - 順序：戶號 | 戶名 | 面積（坪）"""
        self.household_table.setRowCount(0)
        
        for household in households:
            row_position = self.household_table.rowCount()
            self.household_table.insertRow(row_position)
            
            # 戶號
            self.household_table.setItem(
                row_position, 0,
                QTableWidgetItem(household['household_id'])
            )
            # 戶名
            self.household_table.setItem(
                row_position, 1,
                QTableWidgetItem(household['name'])
            )
            # 面積（坪）
            share_amount = household.get('share_amount', 0.0)
            share_item = QTableWidgetItem(str(share_amount))
            share_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.household_table.setItem(row_position, 2, share_item)
    
    def filter_households(self):
        """根據搜尋條件過濾住戶"""
        search_text = self.search_input.text().strip().lower()
        
        if not search_text:
            self.refresh_table(self.all_households)
            return
        
        filtered = [
            h for h in self.all_households
            if search_text in h['household_id'].lower() or
               search_text in h['name'].lower()
        ]
        
        self.refresh_table(filtered)
    
    def _parse_share_amount(self, value) -> float:
        """解析面積數值"""
        if not value:
            return 0.0
        
        try:
            return float(str(value).strip())
        except (ValueError, TypeError):
            return 0.0
    
    def import_households(self):
        """從 .xlsx/.csv 文件導入住戶"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "選擇住戶文件",
            "",
            "Excel 文件 (*.xlsx);;CSV 文件 (*.csv);;所有文件 (*.*)"
        )
        
        if not file_path:
            return
        
        try:
            if file_path.endswith('.xlsx'):
                households = self._read_xlsx_file(file_path)
            else:
                households = self._read_csv_file(file_path)
            
            if not households:
                QMessageBox.warning(self, "警告", "文件中沒有有效的住戶數據")
                return
            
            # 導入住戶
            success = 0
            failed = 0
            
            for household in households:
                household_id = household['household_id'].strip()
                name = household['name'].strip()
                share_amount = household.get('share_amount', 0.0)
                
                if not household_id or not name:
                    failed += 1
                    continue
                
                # 先檢查是否已存在
                if self.db.get_household(household_id):
                    failed += 1
                    continue
                
                # 添加住戶
                if self.db.add_household(household_id, name, share_amount):
                    success += 1
                else:
                    failed += 1
            
            QMessageBox.information(
                self, "導入完成",
                f"成功導入 {success} 個住戶\n"
                f"失敗 {failed} 個（可能是戶號重複或格式錯誤）"
            )
            
            self.load_households()
            self.search_input.clear()
            
        except Exception as e:
            QMessageBox.critical(self, "錯誤", f"導入失敗: {str(e)}")
    
    def _read_csv_file(self, file_path: str) -> List[Dict]:
        """
        讀取 CSV 文件
        
        期望的列：戶號 | 戶名 | 面積（坪）
        或任何包含這些關鍵字的列名
        """
        households = []
        
        try:
            with open(file_path, 'r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                
                if not reader.fieldnames:
                    raise ValueError("CSV 文件為空")
                
                # 找到相關列（支持多種列名）
                fieldnames_lower = [h.lower().strip() for h in reader.fieldnames]
                
                household_id_col = None
                name_col = None
                share_col = None
                
                for idx, field in enumerate(fieldnames_lower):
                    if '戶號' in field or 'household' in field:
                        household_id_col = reader.fieldnames[idx]
                    elif '姓名' in field or '名稱' in field or '戶名' in field or 'name' in field:
                        name_col = reader.fieldnames[idx]
                    elif '面積' in field or '坪' in field or 'area' in field:
                        share_col = reader.fieldnames[idx]
                
                if not household_id_col:
                    raise ValueError("找不到 '戶號' 列")
                if not name_col:
                    raise ValueError("找不到 '姓名/戶名' 列")
                
                # 讀取數據
                for row in reader:
                    household_id = row.get(household_id_col, '').strip()
                    name = row.get(name_col, '').strip()
                    share_amount = self._parse_share_amount(row.get(share_col, 0.0))
                    
                    if household_id and name:
                        households.append({
                            'household_id': household_id,
                            'name': name,
                            'share_amount': share_amount,
                        })
        
        except Exception as e:
            raise ValueError(f"CSV 讀取失敗: {str(e)}")
        
        return households
    
    def _read_xlsx_file(self, file_path: str) -> List[Dict]:
        """
        讀取 .xlsx 文件
        
        期望的列：戶號 | 戶名 | 面積（坪）
        或任何包含這些關鍵字的列名
        """
        households = []
        
        if XLSX_AVAILABLE:
            from openpyxl import load_workbook
            
            try:
                workbook = load_workbook(file_path)
                worksheet = workbook.active
                
                # 讀取標題行
                headers = []
                for cell in worksheet[1]:
                    headers.append(cell.value)
                
                # 找到相關列
                household_id_col = None
                name_col = None
                share_col = None
                
                for idx, header in enumerate(headers):
                    if not header:
                        continue
                    header_lower = str(header).lower().strip()
                    
                    if '戶號' in header_lower or 'household' in header_lower:
                        household_id_col = idx
                    elif '姓名' in header_lower or '名稱' in header_lower or '戶名' in header_lower or 'name' in header_lower:
                        name_col = idx
                    elif '面積' in header_lower or '坪' in header_lower or 'area' in header_lower:
                        share_col = idx
                
                if household_id_col is None:
                    raise ValueError("找不到 '戶號' 列")
                if name_col is None:
                    raise ValueError("找不到 '姓名/戶名' 列")
                
                # 讀取數據行
                for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                    if not row[household_id_col]:
                        continue
                    
                    household_id = str(row[household_id_col]).strip()
                    name = str(row[name_col]).strip() if row[name_col] else ''
                    share_amount = self._parse_share_amount(row[share_col]) if share_col is not None else 0.0
                    
                    if household_id and name:
                        households.append({
                            'household_id': household_id,
                            'name': name,
                            'share_amount': share_amount,
                        })
            
            except Exception as e:
                raise ValueError(f"XLSX 讀取失敗: {str(e)}")
        
        elif PANDAS_AVAILABLE:
            import pandas as pd
            
            try:
                df = pd.read_excel(file_path)
                
                # 找到相關列
                household_id_col = None
                name_col = None
                share_col = None
                
                for col in df.columns:
                    col_lower = str(col).lower().strip()
                    
                    if '戶號' in col_lower or 'household' in col_lower:
                        household_id_col = col
                    elif '姓名' in col_lower or '名稱' in col_lower or '戶名' in col_lower or 'name' in col_lower:
                        name_col = col
                    elif '面積' in col_lower or '坪' in col_lower or 'area' in col_lower:
                        share_col = col
                
                if not household_id_col:
                    raise ValueError("找不到 '戶號' 列")
                if not name_col:
                    raise ValueError("找不到 '姓名/戶名' 列")
                
                # 讀取數據
                for idx, row in df.iterrows():
                    household_id = str(row[household_id_col]).strip() if pd.notna(row[household_id_col]) else ''
                    name = str(row[name_col]).strip() if pd.notna(row[name_col]) else ''
                    share_amount = self._parse_share_amount(row[share_col]) if share_col and pd.notna(row[share_col]) else 0.0
                    
                    if household_id and household_id != 'nan' and name and name != 'nan':
                        households.append({
                            'household_id': household_id,
                            'name': name,
                            'share_amount': share_amount,
                        })
            
            except Exception as e:
                raise ValueError(f"XLSX 讀取失敗: {str(e)}")
        
        return households
    
    def export_check_in_data(self):
        """導出報到資料到 .xlsx 文件（可選擇目錄和文件名）"""
        if not XLSX_AVAILABLE and not PANDAS_AVAILABLE:
            QMessageBox.critical(
                self, "錯誤",
                "需要安裝 openpyxl 或 pandas\n\n"
                "請運行: pip install openpyxl pandas"
            )
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "保存報到資料",
            "報到資料.xlsx",
            "Excel 文件 (*.xlsx)"
        )
        
        if not file_path:
            return
        
        try:
            # 獲取所有住戶數據
            households_data = []
            
            for household in self.all_households:
                households_data.append({
                    '戶號': household['household_id'],
                    '戶名': household['name'],
                    '面積（坪）': household.get('share_amount', 0.0),
                })
            
            if not households_data:
                QMessageBox.warning(self, "警告", "沒有住戶數據可導出")
                return
            
            if PANDAS_AVAILABLE:
                # 使用 pandas 導出
                import pandas as pd
                
                df = pd.DataFrame(households_data)
                df.to_excel(file_path, index=False, sheet_name='報到資料')
            
            elif XLSX_AVAILABLE:
                # 使用 openpyxl 導出
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
                
                workbook = Workbook()
                worksheet = workbook.active
                worksheet.title = '報到資料'
                
                # 設置列寬
                worksheet.column_dimensions['A'].width = 12
                worksheet.column_dimensions['B'].width = 18
                worksheet.column_dimensions['C'].width = 15
                
                # 寫入標題
                headers = ['戶號', '戶名', '面積（坪）']
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(bold=True, size=11, color="FFFFFF")
                
                for col_idx, header in enumerate(headers, start=1):
                    cell = worksheet.cell(row=1, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # 寫入數據
                for row_idx, household in enumerate(households_data, start=2):
                    worksheet.cell(row=row_idx, column=1, value=household['戶號'])
                    worksheet.cell(row=row_idx, column=2, value=household['戶名'])
                    worksheet.cell(row=row_idx, column=3, value=household['面積（坪）'])
                
                workbook.save(file_path)
            
            QMessageBox.information(
                self, "成功",
                f"已導出 {len(households_data)} 個住戶到\n{Path(file_path).name}"
            )
            
        except Exception as e:
            QMessageBox.critical(self, "錯誤", f"導出失敗: {str(e)}")

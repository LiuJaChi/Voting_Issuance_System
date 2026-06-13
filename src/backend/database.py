"""
SQLite 數據庫管理模塊
"""
import json
import sqlite3
from pathlib import Path
from typing import List, Dict, Optional, Tuple
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import platform


class Database:
    """SQLite 數據庫管理類"""

    def __init__(self, db_path: str = "data/votes.db", config_manager=None):
        """初始化數據庫連接
        
        Args:
            db_path: 數據庫文件路徑
            config_manager: 配置管理器實例，用於獲取系統配置（預期人數等）
        """
        self.db_path = db_path
        self.config_manager = config_manager
        Path(db_path).parent.mkdir(parents=True, exist_ok=True)
        self.init_db()
        
        # 初始化中文字體支持
        self._setup_chinese_font()

    def _setup_chinese_font(self):
        """設置中文字體支持"""
        try:
            system = platform.system()
            font_path = None
            
            # 根據系統選擇字體
            if system == 'Windows':
                # Windows 系統使用微軟雅黑或新細明體
                possible_paths = [
                    'C:\\Windows\\Fonts\\msyh.ttc',      # 微軟雅黑
                    'C:\\Windows\\Fonts\\msjh.ttc',      # 微軟正黑
                    'C:\\Windows\\Fonts\\SimSun.ttc',    # 宋體
                ]
                for path in possible_paths:
                    if Path(path).exists():
                        font_path = path
                        break
            elif system == 'Darwin':  # macOS
                possible_paths = [
                    '/Library/Fonts/Microsoft YaHei.ttf',
                    '/System/Library/Fonts/STHeiti Light.ttc',
                    '/System/Library/Fonts/PingFang.ttc',
                ]
                for path in possible_paths:
                    if Path(path).exists():
                        font_path = path
                        break
            else:  # Linux
                possible_paths = [
                    '/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc',
                    '/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc',
                    '/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc',
                ]
                for path in possible_paths:
                    if Path(path).exists():
                        font_path = path
                        break
            
            # 如果找到字體，註冊到 ReportLab
            if font_path:
                pdfmetrics.registerFont(TTFont('ChineseFont', font_path))
                self.chinese_font = 'ChineseFont'
            else:
                # 如果找不到字體，使用默認字體（不支持中文，但不會崩潰）
                self.chinese_font = 'Helvetica'
                print("⚠ 警告：找不到中文字體，PDF 中文可能無法顯示")
        except Exception as e:
            print(f"⚠ 字體設置失敗: {e}")
            self.chinese_font = 'Helvetica'

    def get_connection(self):
        """獲取數據庫連接"""
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        return conn

    def init_db(self):
        """初始化數據庫表"""
        conn = self.get_connection()
        cursor = conn.cursor()

        # 住戶表
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS households (
                household_id TEXT PRIMARY KEY,
                name TEXT NOT NULL,
                share_amount REAL DEFAULT 0.0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)

        # 報到記錄表
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS check_in_records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                household_id TEXT NOT NULL UNIQUE,
                checked_in_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                device_id TEXT,
                FOREIGN KEY (household_id) REFERENCES households(household_id)
            )
        """)

        # 投票項目表
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS voting_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                case_number TEXT UNIQUE NOT NULL,
                name TEXT NOT NULL,
                description TEXT,
                vote_type TEXT DEFAULT '一般議案',
                pass_percentage REAL DEFAULT 66.7,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)

        # 投票記錄表（複合主鍵：household_id + case_number）
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS votes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                household_id TEXT NOT NULL,
                case_number TEXT NOT NULL,
                vote TEXT NOT NULL,
                device_id TEXT,
                voted_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(household_id, case_number),
                FOREIGN KEY (household_id) REFERENCES households(household_id),
                FOREIGN KEY (case_number) REFERENCES voting_items(case_number)
            )
        """)

        # 條碼映射表
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS barcode_mapping (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                household_id TEXT NOT NULL UNIQUE,
                barcode_data TEXT NOT NULL UNIQUE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (household_id) REFERENCES households(household_id)
            )
        """)

        conn.commit()
        conn.close()

    # ─────────────────────────── 住戶管理 ───────────────────────────

    def add_household(self, household_id: str, name: str, share_amount: float = 0.0) -> bool:
        """添加住戶"""
        conn = self.get_connection()
        cursor = conn.cursor()
        try:
            cursor.execute("""
                INSERT INTO households (household_id, name, share_amount)
                VALUES (?, ?, ?)
            """, (household_id, name, share_amount))
            conn.commit()
            conn.close()
            return True
        except sqlite3.IntegrityError:
            conn.close()
            return False

    def update_household(self, household_id: str, name: str, share_amount: float = None) -> bool:
        """更新住戶信息"""
        conn = self.get_connection()
        cursor = conn.cursor()

        if share_amount is not None:
            cursor.execute("""
                UPDATE households SET name = ?, share_amount = ? WHERE household_id = ?
            """, (name, share_amount, household_id))
        else:
            cursor.execute("""
                UPDATE households SET name = ? WHERE household_id = ?
            """, (name, household_id))

        conn.commit()
        conn.close()
        return True

    def delete_household(self, household_id: str) -> bool:
        """刪除住戶"""
        conn = self.get_connection()
        cursor = conn.cursor()
        try:
            cursor.execute("DELETE FROM households WHERE household_id = ?", (household_id,))
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            print(f"刪除住戶失敗: {e}")
            conn.close()
            return False

    def get_household(self, household_id: str) -> Optional[Dict]:
        """獲取住戶信息"""
        conn = self.get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM households WHERE household_id = ?", (household_id,))
        row = cursor.fetchone()
        conn.close()
        return dict(row) if row else None

    def get_all_households(self) -> List[Dict]:
        """獲取所有住戶"""
        conn = self.get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM households ORDER BY household_id")
        households = [dict(row) for row in cursor.fetchall()]
        conn.close()
        return households

    def import_households(self, households: List[Dict]) -> Tuple[int, int]:
        """批量導入住戶"""
        conn = self.get_connection()
        cursor = conn.cursor()
        success = 0
        skipped = 0

        for household in households:
            try:
                cursor.execute("""
                    INSERT INTO households (household_id, name, share_amount)
                    VALUES (?, ?, ?)
                """, (household['household_id'], household['name'], household.get('share_amount', 0.0)))
                success += 1
            except sqlite3.IntegrityError:
                skipped += 1

        conn.commit()
        conn.close()
        return success, skipped

    # ─────────────────────────── 條碼管理 ───────────────────────────

    def get_household_id_by_barcode(self, barcode_data: str) -> Optional[str]:
        """通過條碼獲取住戶ID"""
        conn = self.get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT household_id FROM barcode_mapping WHERE barcode_data = ?", (barcode_data,))
        row = cursor.fetchone()
        conn.close()
        return row[0] if row else None

    def get_barcode_by_household_id(self, household_id: str) -> Optional[str]:
        """通過住戶ID獲取條碼"""
        conn = self.get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT barcode_data FROM barcode_mapping WHERE household_id = ?", (household_id,))
        row = cursor.fetchone()
        conn.close()
        return row[0] if row else None

    def get_all_barcode_mappings(self) -> List[Dict]:
        """獲取所有條碼映射"""
        conn = self.get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM barcode_mapping")
        mappings = [dict(row) for row in cursor.fetchall()]
        conn.close()
        return mappings

    def delete_barcode_mapping(self, household_id: str) -> bool:
        """刪除條碼映射"""
        conn = self.get_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM barcode_mapping WHERE household_id = ?", (household_id,))
        conn.commit()
        conn.close()
        return True

    # ─────────────────────────── 報到管理 ───────────────────────────

    def add_check_in(self, household_id: str, device_id: str = None) -> bool:
        """添加報到記錄"""
        conn = self.get_connection()
        cursor = conn.cursor()
        try:
            cursor.execute("""
                INSERT INTO check_in_records (household_id, device_id)
                VALUES (?, ?)
            """, (household_id, device_id))
            conn.commit()
            conn.close()
            return True
        except sqlite3.IntegrityError:
            conn.close()
            return False

    def get_check_in_stats(self) -> Dict:
        """
        獲取報到統計
        
        使用配置中的「預期參與人數」而不是實際住戶數
        返回字段：
        - expected_total: 預期出席人數（來自配置）
        - checked_in: 已報到人數
        - not_checked_in: 未報到人數
        - percentage: 報到百分比
        - total_expected: 同 expected_total（向後相容）
        """
        if self.config_manager:
            expected_total = self.config_manager.get_config('total_participants', 0)
        else:
            conn = self.get_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM households")
            expected_total = cursor.fetchone()[0]
            conn.close()
        
        conn = self.get_connection()
        cursor = conn.cursor()

        cursor.execute("SELECT COUNT(*) FROM check_in_records")
        checked_in = cursor.fetchone()[0]

        conn.close()

        not_checked_in = expected_total - checked_in
        percentage = (checked_in / expected_total * 100) if expected_total > 0 else 0

        return {
            'expected_total': expected_total,
            'total_expected': expected_total,  # 向後相容
            'checked_in': checked_in,
            'not_checked_in': not_checked_in,
            'percentage': percentage,
            'checked_in_percentage': percentage  # 向後相容
        }

    def get_check_in_area_stats(self) -> Dict:
        """獲取報到面積統計"""
        conn = self.get_connection()
        cursor = conn.cursor()

        # 總面積
        cursor.execute("SELECT SUM(share_amount) FROM households")
        total_area = cursor.fetchone()[0] or 0.0

        # 已報到面積
        cursor.execute("""
            SELECT SUM(h.share_amount)
            FROM households h
            LEFT JOIN check_in_records c ON h.household_id = c.household_id
            WHERE c.household_id IS NOT NULL
        """)
        checked_in_area = cursor.fetchone()[0] or 0.0

        conn.close()

        return {
            'total_area': total_area,
            'checked_in_area': checked_in_area,
            'not_checked_in_area': total_area - checked_in_area,
            'checked_in_area_percentage': (checked_in_area / total_area * 100) if total_area > 0 else 0
        }

    def is_checked_in(self, household_id: str) -> bool:
        """檢查是否已報到"""
        conn = self.get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT 1 FROM check_in_records WHERE household_id = ?", (household_id,))
        result = cursor.fetchone() is not None
        conn.close()
        return result

    def clear_check_in_data(self) -> bool:
        """清空報到數據"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            cursor.execute("DELETE FROM check_in_records")
            conn.commit()
            conn.close()
            print("✓ 報到數據已清空")
            return True
        except Exception as e:
            print(f"✗ 清空報到數據失敗: {e}")
            return False

    # ─────────────────────────── 投票項目管理 ───────────────────────────

    def add_voting_item(self, case_number: str, name: str, description: str = '',
                       vote_type: str = '一般議案', pass_percentage: float = 66.7) -> bool:
        """添加投票項目"""
        conn = self.get_connection()
        cursor = conn.cursor()
        try:
            cursor.execute("""
                INSERT INTO voting_items (case_number, name, description, vote_type, pass_percentage)
                VALUES (?, ?, ?, ?, ?)
            """, (case_number, name, description, vote_type, pass_percentage))
            conn.commit()
            conn.close()
            return True
        except sqlite3.IntegrityError:
            conn.close()
            return False

    def update_voting_item(self, case_number: str, name: str, description: str,
                          vote_type: str, pass_percentage: float) -> bool:
        """更新投票項目"""
        conn = self.get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE voting_items
            SET name = ?, description = ?, vote_type = ?, pass_percentage = ?
            WHERE case_number = ?
        """, (name, description, vote_type, pass_percentage, case_number))
        conn.commit()
        conn.close()
        return True

    def delete_voting_item(self, case_number: str) -> bool:
        """刪除投票項目"""
        conn = self.get_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM voting_items WHERE case_number = ?", (case_number,))
        conn.commit()
        conn.close()
        return True

    def get_voting_item(self, case_number: str) -> Optional[Dict]:
        """獲取投票項目"""
        conn = self.get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM voting_items WHERE case_number = ?", (case_number,))
        row = cursor.fetchone()
        conn.close()
        return dict(row) if row else None

    def get_all_voting_items(self) -> List[Dict]:
        """獲取所有投票項目"""
        conn = self.get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM voting_items ORDER BY case_number")
        items = [dict(row) for row in cursor.fetchall()]
        conn.close()
        return items

    # ─────────────────────────── 投票記錄管理 ───────────────────────────

    def add_vote(self, household_id: str, case_number: str, vote: str, device_id: str = None) -> bool:
        """添加投票記錄"""
        conn = self.get_connection()
        cursor = conn.cursor()
        try:
            cursor.execute("""
                INSERT INTO votes (household_id, case_number, vote, device_id)
                VALUES (?, ?, ?, ?)
            """, (household_id, case_number, vote, device_id))
            conn.commit()
            conn.close()
            return True
        except sqlite3.IntegrityError:
            conn.close()
            return False

    def record_vote(self, household_id: str, case_number: str, vote: str, device_id: str = None) -> bool:
        """記錄投票（別名方法）"""
        return self.add_vote(household_id, case_number, vote, device_id)

    def update_vote(self, household_id: str, case_number: str, vote: str) -> bool:
        """更新投票記錄"""
        conn = self.get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE votes SET vote = ? WHERE household_id = ? AND case_number = ?
        """, (vote, household_id, case_number))
        conn.commit()
        conn.close()
        return True

    def delete_vote(self, household_id: str, case_number: str) -> bool:
        """刪除投票記錄"""
        conn = self.get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            DELETE FROM votes WHERE household_id = ? AND case_number = ?
        """, (household_id, case_number))
        conn.commit()
        conn.close()
        return True

    def get_vote(self, household_id: str, case_number: str) -> Optional[str]:
        """獲取投票記錄"""
        conn = self.get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT vote FROM votes WHERE household_id = ? AND case_number = ?
        """, (household_id, case_number))
        row = cursor.fetchone()
        conn.close()
        return row[0] if row else None

    def get_all_votes_for_case(self, case_number: str) -> List[Dict]:
        """獲取某案件的所有投票"""
        conn = self.get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT * FROM votes WHERE case_number = ? ORDER BY household_id
        """, (case_number,))
        votes = [dict(row) for row in cursor.fetchall()]
        conn.close()
        return votes

    def get_voting_data_with_details(self, case_number: str) -> List[Dict]:
        """獲取投票數據（包含住戶詳情）"""
        conn = self.get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT v.*, h.name, h.share_amount
            FROM votes v
            LEFT JOIN households h ON v.household_id = h.household_id
            WHERE v.case_number = ?
        """, (case_number,))
        votes = [dict(row) for row in cursor.fetchall()]
        conn.close()
        return votes

    def get_all_votes_for_household(self, household_id: str) -> List[Dict]:
        """獲取某住戶的所有投票"""
        conn = self.get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT * FROM votes WHERE household_id = ? ORDER BY case_number
        """, (household_id,))
        votes = [dict(row) for row in cursor.fetchall()]
        conn.close()
        return votes

    def has_voted(self, household_id: str, case_number: str) -> bool:
        """檢查是否已投票"""
        conn = self.get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT 1 FROM votes WHERE household_id = ? AND case_number = ?
        """, (household_id, case_number))
        result = cursor.fetchone() is not None
        conn.close()
        return result

    def get_voting_results(self, case_number: str) -> Dict:
        """獲取投票結果統計"""
        conn = self.get_connection()
        cursor = conn.cursor()

        # 獲取投票項目
        voting_item = self.get_voting_item(case_number)
        if not voting_item:
            return {}

        # 計算投票統計
        cursor.execute("""
            SELECT vote, COUNT(*) as count, SUM(h.share_amount) as area
            FROM votes v
            LEFT JOIN households h ON v.household_id = h.household_id
            WHERE v.case_number = ?
            GROUP BY vote
        """, (case_number,))

        results = {
            'case_number': case_number,
            'name': voting_item['name'],
            'by_vote': {},
            'total_count': 0,
            'total_area': 0.0
        }

        for row in cursor.fetchall():
            vote_type = row[0]
            count = row[1]
            area = row[2] or 0.0
            results['by_vote'][vote_type] = {
                'count': count,
                'area': area,
                'area_percentage': 0.0
            }
            results['total_count'] += count
            results['total_area'] += area

        # 計算百分比
        for vote_type in results['by_vote']:
            if results['total_area'] > 0:
                results['by_vote'][vote_type]['area_percentage'] = (
                    results['by_vote'][vote_type]['area'] / results['total_area'] * 100
                )

        conn.close()
        return results

    def get_voting_statistics(self, case_number: str) -> Dict:
        """獲取投票統計（包含完整的統計信息）"""
        conn = self.get_connection()
        cursor = conn.cursor()

        # 基礎結果
        basic_results = self.get_voting_results(case_number)
        
        # 獲取報到的住戶及其面積
        cursor.execute("""
            SELECT COUNT(*) as count, SUM(h.share_amount) as area
            FROM check_in_records c
            LEFT JOIN households h ON c.household_id = h.household_id
        """)
        checked_in_data = cursor.fetchone()
        total_checked_in_households = checked_in_data[0] or 0
        total_checked_in_area = checked_in_data[1] or 0.0
        
        # 該案件的已投票人數和面積
        cursor.execute("""
            SELECT COUNT(*) as count, SUM(h.share_amount) as area
            FROM votes v
            LEFT JOIN households h ON v.household_id = h.household_id
            WHERE v.case_number = ?
        """, (case_number,))
        voted_data = cursor.fetchone()
        total_voted_households = voted_data[0] or 0
        total_voted_area = voted_data[1] or 0.0

        conn.close()

        # 計算各投票選項的人數百分比
        for vote_type, stats in basic_results.get('by_vote', {}).items():
            if total_voted_households > 0:
                stats['count_percentage'] = (stats['count'] / total_voted_households * 100)
            else:
                stats['count_percentage'] = 0.0

        # 計算整體百分比
        overall_count_percentage = (total_voted_households / total_checked_in_households * 100) if total_checked_in_households > 0 else 0
        overall_area_percentage = (total_voted_area / total_checked_in_area * 100) if total_checked_in_area > 0 else 0

        return {
            **basic_results,
            'total_voted_households': total_voted_households,
            'total_voted_area': total_voted_area,
            'total_checked_in_households': total_checked_in_households,
            'total_checked_in_area': total_checked_in_area,
            'overall_count_percentage': overall_count_percentage,
            'overall_area_percentage': overall_area_percentage
        }

    # ─────────────────────────── 數據導出 ───────────────────────────

    def export_voting_data(self, export_path: str = None) -> str:
        """匯出投票數據為 XLSX 文件
        
        Args:
            export_path: 匯出路徑，如果不指定則自動生成
            
        Returns:
            匯出文件路徑，失敗則返回空字符串
        """
        if export_path is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            export_path = f"exports/voting_data_{timestamp}.xlsx"

        try:
            conn = self.get_connection()
            cursor = conn.cursor()

            # 獲取投票數據
            cursor.execute("""
                SELECT v.household_id, v.case_number, v.vote, h.name, h.share_amount, v.voted_at
                FROM votes v
                LEFT JOIN households h ON v.household_id = h.household_id
                ORDER BY v.case_number, v.household_id
            """)
            votes = cursor.fetchall()

            # 獲取投票項目
            cursor.execute("SELECT * FROM voting_items ORDER BY case_number")
            voting_items = cursor.fetchall()

            conn.close()

            # 創建 Excel 工作簿
            wb = Workbook()
            ws_votes = wb.active
            ws_votes.title = "投票記錄"

            # 投票記錄工作表
            ws_votes.append(["戶號", "案號", "投票", "戶名", "面積(坪)", "投票時間"])
            for vote in votes:
                ws_votes.append([
                    vote[0],  # household_id
                    vote[1],  # case_number
                    vote[2],  # vote
                    vote[3],  # name
                    f"{vote[4]:.2f}" if vote[4] else "0.00",  # share_amount
                    vote[5]   # voted_at
                ])

            # 投票項目工作表
            ws_items = wb.create_sheet("投票項目")
            ws_items.append(["案號", "項目名稱", "描述", "投票類型", "通過百分比"])
            for item in voting_items:
                ws_items.append([
                    item[1],  # case_number
                    item[2],  # name
                    item[3],  # description
                    item[4],  # vote_type
                    item[5]   # pass_percentage
                ])

            # 創建導出目錄
            Path(export_path).parent.mkdir(parents=True, exist_ok=True)

            # 保存工作簿
            wb.save(export_path)
            print(f"✓ 投票數據已匯出到: {export_path}")
            return export_path

        except Exception as e:
            print(f"✗ 投票數據匯出失敗: {e}")
            return ""

    def import_voting_data(self, file_path: str, mode: str = 'merge') -> Dict:
        """匯入投票數據
        
        Args:
            file_path: XLSX 文件路徑
            mode: 匯入模式 - 'replace' 替換所有 / 'merge' 合併
            
        Returns:
            包含 messages 和 errors 的字典
        """
        result = {'messages': [], 'errors': []}
        
        try:
            wb = load_workbook(file_path)
            
            # 檢查工作表
            if '投票記錄' not in wb.sheetnames:
                result['errors'].append("找不到 '投票記錄' 工作表")
                return result
            
            ws_votes = wb['投票記錄']
            
            # 替換模式：清空現有投票數據
            if mode == 'replace':
                conn = self.get_connection()
                cursor = conn.cursor()
                cursor.execute("DELETE FROM votes")
                conn.commit()
                conn.close()
            
            # 匯入投票數據
            conn = self.get_connection()
            cursor = conn.cursor()
            
            imported_count = 0
            skipped_count = 0
            
            for row_idx, row in enumerate(ws_votes.iter_rows(min_row=2, values_only=True), start=2):
                if not row[0]:  # 跳過空行
                    continue
                
                try:
                    household_id = str(row[0]).strip()
                    case_number = str(row[1]).strip()
                    vote = str(row[2]).strip()
                    
                    # 驗證投票選項
                    if vote not in ['同意', '不同意', '棄權']:
                        result['errors'].append(f"第 {row_idx} 行: 無效的投票選項 '{vote}'")
                        skipped_count += 1
                        continue
                    
                    # 檢查住戶是否存在
                    if not self.get_household(household_id):
                        result['errors'].append(f"第 {row_idx} 行: 住戶 '{household_id}' 不存在")
                        skipped_count += 1
                        continue
                    
                    # 檢查投票項目是否存在
                    if not self.get_voting_item(case_number):
                        result['errors'].append(f"第 {row_idx} 行: 案號 '{case_number}' 不存在")
                        skipped_count += 1
                        continue
                    
                    # 嘗試插入或更新
                    try:
                        cursor.execute("""
                            INSERT INTO votes (household_id, case_number, vote)
                            VALUES (?, ?, ?)
                        """, (household_id, case_number, vote))
                        imported_count += 1
                    except sqlite3.IntegrityError:
                        # 如果是合併模式，更新現有記錄
                        if mode == 'merge':
                            cursor.execute("""
                                UPDATE votes SET vote = ? 
                                WHERE household_id = ? AND case_number = ?
                            """, (vote, household_id, case_number))
                            imported_count += 1
                        else:
                            skipped_count += 1
                
                except Exception as e:
                    result['errors'].append(f"第 {row_idx} 行: {str(e)}")
                    skipped_count += 1
            
            conn.commit()
            conn.close()
            
            result['messages'].append(f"✓ 成功匯入 {imported_count} 筆投票記錄")
            if skipped_count > 0:
                result['messages'].append(f"⚠ 略過 {skipped_count} 筆記錄")
            
            return result
        
        except Exception as e:
            result['errors'].append(f"匯入失敗: {str(e)}")
            return result

    def export_data(self, export_path: str = None) -> bool:
        """導出所有數據到 JSON 文件"""
        if export_path is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            export_path = f"exports/data_{timestamp}.json"

        try:
            conn = self.get_connection()
            cursor = conn.cursor()

            # 獲取所有表的數據
            cursor.execute("SELECT * FROM households")
            households = [dict(row) for row in cursor.fetchall()]

            cursor.execute("SELECT * FROM check_in_records")
            check_in_records = [dict(row) for row in cursor.fetchall()]

            cursor.execute("SELECT * FROM barcode_mapping")
            barcode_mappings = [dict(row) for row in cursor.fetchall()]

            cursor.execute("SELECT * FROM voting_items")
            voting_items = [dict(row) for row in cursor.fetchall()]

            cursor.execute("SELECT * FROM votes")
            votes = [dict(row) for row in cursor.fetchall()]

            conn.close()

            # 構建導出數據
            export_data = {
                'config': {
                    'export_time': datetime.now().isoformat()
                },
                'households': households,
                'check_in_records': check_in_records,
                'barcode_mappings': barcode_mappings,
                'voting_items': voting_items,
                'votes': votes
            }

            # 創建導出目錄
            Path(export_path).parent.mkdir(parents=True, exist_ok=True)

            # 保存 JSON 文件
            with open(export_path, 'w', encoding='utf-8') as f:
                json.dump(export_data, f, ensure_ascii=False, indent=2, default=str)

            print(f"✓ 數據已導出到: {export_path}")
            return True
        except Exception as e:
            print(f"✗ 數據導出失敗: {e}")
            return False

    def import_json_data(self, file_path: str) -> bool:
        """從 JSON 文件導入所有數據（由 export_data() 導出的格式）

        Args:
            file_path: JSON 文件路徑

        Returns:
            導入成功返回 True，失敗返回 False
        """
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)

            conn = self.get_connection()
            cursor = conn.cursor()

            # 導入住戶數據
            for household in data.get('households', []):
                cursor.execute("""
                    INSERT OR REPLACE INTO households
                        (household_id, name, share_amount, created_at)
                    VALUES (:household_id, :name, :share_amount, :created_at)
                """, household)

            # 導入報到記錄
            for record in data.get('check_in_records', []):
                cursor.execute("""
                    INSERT OR REPLACE INTO check_in_records
                        (household_id, checked_in_at, device_id)
                    VALUES (:household_id, :checked_in_at, :device_id)
                """, record)

            # 導入條碼映射
            for mapping in data.get('barcode_mappings', []):
                cursor.execute("""
                    INSERT OR REPLACE INTO barcode_mapping
                        (household_id, barcode_data, created_at)
                    VALUES (:household_id, :barcode_data, :created_at)
                """, mapping)

            # 導入投票項目
            for item in data.get('voting_items', []):
                cursor.execute("""
                    INSERT OR REPLACE INTO voting_items
                        (case_number, name, description, vote_type, pass_percentage, created_at)
                    VALUES (:case_number, :name, :description, :vote_type, :pass_percentage, :created_at)
                """, item)

            # 導入投票記錄
            for vote in data.get('votes', []):
                cursor.execute("""
                    INSERT OR REPLACE INTO votes
                        (household_id, case_number, vote, device_id, voted_at)
                    VALUES (:household_id, :case_number, :vote, :device_id, :voted_at)
                """, vote)

            conn.commit()
            conn.close()

            print(f"✓ 數據已從 {file_path} 導入")
            return True
        except Exception as e:
            print(f"✗ 數據導入失敗: {e}")
            return False

    def export_voting_results_pdf(self, export_path: str = None) -> str:
        """匯出投票結果為 PDF（支持繁體中文）
        
        Args:
            export_path: 匯出路徑，如果不指定則自動生成
            
        Returns:
            匯出文件路徑，失敗則返回空字符串
        """
        if export_path is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            export_path = f"exports/voting_results_{timestamp}.pdf"

        try:
            Path(export_path).parent.mkdir(parents=True, exist_ok=True)

            # 建立 PDF 文檔（橫向 A4，以容納 12 欄）
            doc = SimpleDocTemplate(
                export_path,
                pagesize=landscape(A4),
                topMargin=15*mm, bottomMargin=15*mm,
                leftMargin=10*mm, rightMargin=10*mm
            )
            story = []

            # 設置樣式
            styles = getSampleStyleSheet()

            # 標題樣式：白色字體，藍色底色
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontName=self.chinese_font,
                fontSize=16,
                textColor=colors.white,
                backColor=colors.HexColor('#1976D2'),
                spaceAfter=6,
                spaceBefore=4,
                alignment=TA_CENTER
            )

            # 時間戳樣式
            timestamp_style = ParagraphStyle(
                'Timestamp',
                parent=styles['Normal'],
                fontName=self.chinese_font,
                fontSize=9,
                textColor=colors.grey,
                alignment=TA_CENTER,
                spaceAfter=12
            )

            # 表格標題列樣式（白色字體）
            header_style = ParagraphStyle(
                'TableHeader',
                parent=styles['Normal'],
                fontName=self.chinese_font,
                fontSize=9,
                textColor=colors.white,
                alignment=TA_CENTER
            )

            # 表格資料列樣式
            normal_style = ParagraphStyle(
                'TableNormal',
                parent=styles['Normal'],
                fontName=self.chinese_font,
                fontSize=9,
                alignment=TA_CENTER
            )

            # 標題
            story.append(Paragraph("投票結果統計報告", title_style))
            story.append(Spacer(1, 6))

            # 時間戳
            story.append(Paragraph(
                f"生成時間: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                timestamp_style
            ))
            story.append(Spacer(1, 8))

            # 獲取所有投票項目
            voting_items = self.get_all_voting_items()

            if not voting_items:
                story.append(Paragraph("沒有投票項目資料", normal_style))
            else:
                # 表格標題列（與 UI 欄位一致）
                table_data = [
                    [
                        Paragraph('案號', header_style),
                        Paragraph('項目名稱', header_style),
                        Paragraph('同意', header_style),
                        Paragraph('不同意', header_style),
                        Paragraph('棄權', header_style),
                        Paragraph('同意坪', header_style),
                        Paragraph('不同意坪', header_style),
                        Paragraph('棄權坪', header_style),
                        Paragraph('同意%', header_style),
                        Paragraph('不同意%', header_style),
                        Paragraph('棄權%', header_style),
                        Paragraph('進度', header_style),
                    ]
                ]

                for item in voting_items:
                    stats = self.get_voting_statistics(item['case_number'])
                    vote_stats = stats.get('by_vote', {}) if stats else {}

                    agree_stats = vote_stats.get('同意', {})
                    disagree_stats = vote_stats.get('不同意', {})
                    abstain_stats = vote_stats.get('棄權', {})

                    agree_count = agree_stats.get('count', 0)
                    disagree_count = disagree_stats.get('count', 0)
                    abstain_count = abstain_stats.get('count', 0)

                    agree_area = agree_stats.get('area', 0.0)
                    disagree_area = disagree_stats.get('area', 0.0)
                    abstain_area = abstain_stats.get('area', 0.0)

                    agree_area_pct = agree_stats.get('area_percentage', 0.0)
                    disagree_area_pct = disagree_stats.get('area_percentage', 0.0)
                    abstain_area_pct = abstain_stats.get('area_percentage', 0.0)
                    case_progress = stats.get('overall_count_percentage', 0.0) if stats else 0.0

                    table_data.append([
                        Paragraph(item['case_number'], normal_style),
                        Paragraph(item['name'], normal_style),
                        Paragraph(str(agree_count), normal_style),
                        Paragraph(str(disagree_count), normal_style),
                        Paragraph(str(abstain_count), normal_style),
                        Paragraph(f"{agree_area:.2f}", normal_style),
                        Paragraph(f"{disagree_area:.2f}", normal_style),
                        Paragraph(f"{abstain_area:.2f}", normal_style),
                        Paragraph(f"{agree_area_pct:.2f}%", normal_style),
                        Paragraph(f"{disagree_area_pct:.2f}%", normal_style),
                        Paragraph(f"{abstain_area_pct:.2f}%", normal_style),
                        Paragraph(f"{case_progress:.2f}%", normal_style),
                    ])

                # 欄寬（橫向 A4 可用寬約 247mm）
                # 案號, 項目名稱, 同意, 不同意, 棄權, 同意坪, 不同意坪, 棄權坪, 同意%, 不同意%, 棄權%, 進度
                col_widths = [22*mm, 70*mm, 18*mm, 22*mm, 18*mm,
                              22*mm, 25*mm, 22*mm, 20*mm, 22*mm, 20*mm, 20*mm]

                # 欄索引常數
                COL_AGREE       = 2   # 同意（人數）
                COL_DISAGREE    = 3   # 不同意（人數）
                COL_ABSTAIN     = 4   # 棄權（人數）
                COL_AGREE_AREA  = 5   # 同意坪
                COL_DIS_AREA    = 6   # 不同意坪
                COL_ABS_AREA    = 7   # 棄權坪
                COL_AGREE_PCT   = 8   # 同意%
                COL_DIS_PCT     = 9   # 不同意%
                COL_ABS_PCT     = 10  # 棄權%

                table = Table(table_data, colWidths=col_widths)
                num_rows = len(table_data)
                table.setStyle(TableStyle([
                    # 標題列：藍底白字
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1976D2')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('FONTNAME', (0, 0), (-1, 0), self.chinese_font),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                    ('TOPPADDING', (0, 0), (-1, 0), 6),
                    # 資料列字體
                    ('FONTNAME', (0, 1), (-1, -1), self.chinese_font),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    # 格線
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    # 資料列交替底色
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
                    # 同意欄（綠色）
                    ('BACKGROUND', (COL_AGREE,      1), (COL_AGREE,      num_rows - 1), colors.HexColor('#C8E6C9')),
                    ('BACKGROUND', (COL_AGREE_AREA, 1), (COL_AGREE_AREA, num_rows - 1), colors.HexColor('#C8E6C9')),
                    ('BACKGROUND', (COL_AGREE_PCT,  1), (COL_AGREE_PCT,  num_rows - 1), colors.HexColor('#C8E6C9')),
                    # 不同意欄（紅色）
                    ('BACKGROUND', (COL_DISAGREE, 1), (COL_DISAGREE, num_rows - 1), colors.HexColor('#FFCDD2')),
                    ('BACKGROUND', (COL_DIS_AREA, 1), (COL_DIS_AREA, num_rows - 1), colors.HexColor('#FFCDD2')),
                    ('BACKGROUND', (COL_DIS_PCT,  1), (COL_DIS_PCT,  num_rows - 1), colors.HexColor('#FFCDD2')),
                    # 棄權欄（黃色）
                    ('BACKGROUND', (COL_ABSTAIN,  1), (COL_ABSTAIN,  num_rows - 1), colors.HexColor('#FFF9C4')),
                    ('BACKGROUND', (COL_ABS_AREA, 1), (COL_ABS_AREA, num_rows - 1), colors.HexColor('#FFF9C4')),
                    ('BACKGROUND', (COL_ABS_PCT,  1), (COL_ABS_PCT,  num_rows - 1), colors.HexColor('#FFF9C4')),
                ]))

                story.append(table)

            # 構建 PDF
            doc.build(story)
            print(f"✓ 投票結果 PDF 已匯出到: {export_path}")
            return export_path

        except Exception as e:
            print(f"✗ PDF 匯出失敗: {e}")
            return ""

    # ─────────────────────────── 數據清空 ───────────────────────────

    def clear_all_data(self) -> bool:
        """清空所有數據（所有表）"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()

            # 刪除所有表中的數據
            cursor.execute("DELETE FROM votes")
            cursor.execute("DELETE FROM check_in_records")
            cursor.execute("DELETE FROM barcode_mapping")
            cursor.execute("DELETE FROM voting_items")
            cursor.execute("DELETE FROM households")

            conn.commit()
            conn.close()

            print("✓ 所有數據已清空")
            return True
        except Exception as e:
            print(f"✗ 清空數據失敗: {e}")
            return False

    def clear_household_data(self) -> bool:
        """清空住戶及相關數據（保留投票項目）"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()

            cursor.execute("DELETE FROM votes")
            cursor.execute("DELETE FROM check_in_records")
            cursor.execute("DELETE FROM barcode_mapping")
            cursor.execute("DELETE FROM households")

            conn.commit()
            conn.close()

            print("✓ 住戶數據已清空")
            return True
        except Exception as e:
            print(f"✗ 清空住戶數據失敗: {e}")
            return False

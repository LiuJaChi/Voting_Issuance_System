"""
報到單 PDF 生成模塊 + 報到條碼 Excel 導出 - 使用 QR Code 生成報到標籤

報到單規格：
- 每張標籤：戶號 + QR Code 圖像
- 每張大小：BARCODE 標籤尺寸（90mm × 35mm）
- 每頁 A4：2 欄 × 8 列 = 最多 16 張
- 內容：戶號（上方） + QR Code 圖像（下方）

報到.xlsx 導出欄位：戶號 | 戶名 | 面積（坪） | 條碼
"""
import os
import tempfile
from pathlib import Path
from typing import List, Dict
import json

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.platypus import Image as RLImage
from reportlab.platypus import Spacer
import qrcode

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False


# 報到單（標籤）尺寸
LABEL_WIDTH = 90 * mm
LABEL_HEIGHT = 35 * mm

# 每頁欄數與列數
COLS_PER_PAGE = 2
ROWS_PER_PAGE = 8
QR_BOX_SIZE = 8
QR_BORDER = 2


class CheckInPrinter:
    """報到單 PDF 生成器 + 報到條碼 Excel 導出"""

    def __init__(self, output_dir: str = "exports/check_in_ballots"):
        """初始化"""
        self.output_dir = output_dir
        Path(output_dir).mkdir(parents=True, exist_ok=True)
        self.temp_barcodes = []  # 存儲臨時 QR Code 文件路徑

    def _cleanup_temp_files(self):
        """清理臨時 QR Code 文件"""
        for temp_path in self.temp_barcodes:
            try:
                if os.path.exists(temp_path):
                    os.remove(temp_path)
            except Exception as e:
                print(f"清理臨時文件失敗 {temp_path}: {e}")
        self.temp_barcodes = []

    def _build_qr_payload(self, household: Dict) -> str:
        """構建 QR Code JSON 內容。

        欄位包含 household_id、name、share_amount，
        供手機掃描端與後續統計擴充使用。
        """
        household_id = household.get('household_id', '').strip()
        if not household_id:
            raise ValueError('household_id is required for QR payload')

        payload = {
            'household_id': household_id,
            'name': household.get('name', ''),
            'share_amount': household.get('share_amount', 0.0),
        }
        return json.dumps(payload, ensure_ascii=False, separators=(',', ':'))

    def _generate_qr_code_image(self, household: Dict) -> str:
        """生成 QR Code 圖片，返回文件路徑"""
        household_id = household['household_id']
        temp_dir = tempfile.gettempdir()
        safe_content = household_id.replace('/', '_').replace('\\', '_')
        temp_path = os.path.join(temp_dir, f"qrcode_{safe_content}.png")

        qr = qrcode.QRCode(
            version=None,
            error_correction=qrcode.constants.ERROR_CORRECT_M,
            box_size=QR_BOX_SIZE,
            border=QR_BORDER,
        )
        qr.add_data(self._build_qr_payload(household))
        qr.make(fit=True)

        qr_img = qr.make_image(fill_color='black', back_color='white')
        qr_img.save(temp_path)

        self.temp_barcodes.append(temp_path)
        return temp_path

    def generate_pdf(
        self,
        households: List[Dict],
        filename: str = "check_in_ballots.pdf"
    ) -> str:
        """
        生成報到單 PDF
        
        報到單格式：
        戶號（上方） + QR Code 圖像（下方）
        
        Args:
            households: [{'household_id': 'A106-02', 'name': '洪正平'}, ...]
            filename: 輸出文件名
            
        Returns:
            輸出文件路徑
        """
        output_path = str(Path(self.output_dir) / filename)

        page_w, page_h = A4
        margin = 10 * mm

        doc = SimpleDocTemplate(
            output_path,
            pagesize=A4,
            leftMargin=margin,
            rightMargin=margin,
            topMargin=margin,
            bottomMargin=margin,
        )

        styles = getSampleStyleSheet()
        
        # 戶號樣式（加粗、居中）
        household_id_style = ParagraphStyle(
            'HouseholdID',
            parent=styles['Normal'],
            alignment=TA_CENTER,
            fontSize=12,
            leading=14,
            fontName='Helvetica-Bold',
        )

        # 每個標籤的寬度
        available_w = page_w - 2 * margin
        cell_w = available_w / COLS_PER_PAGE

        # 每個標籤的高度
        cell_h = LABEL_HEIGHT

        table_data = []
        row_cells = []

        for household in households:
            household_id = household['household_id']

            # 生成 QR Code 圖像
            try:
                qr_code_path = self._generate_qr_code_image(household)
                qr_code_img = RLImage(qr_code_path, width=18 * mm, height=18 * mm)
            except Exception as e:
                print(f"QR Code 生成失敗 {household_id}: {e}")
                qr_code_img = Paragraph(f"Error: {household_id}", household_id_style)

            # 單元格內容：戶號（上方）+ QR Code 圖像（下方）
            # 使用垂直排列
            cell_content_table = Table(
                [
                    [Paragraph(household_id, household_id_style)],
                    [Spacer(1, 2 * mm)],  # 間距
                    [qr_code_img],
                ],
                colWidths=[cell_w * 0.95],
                rowHeights=[8 * mm, 2 * mm, 18 * mm],
            )
            cell_content_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))

            row_cells.append(cell_content_table)

            if len(row_cells) == COLS_PER_PAGE:
                table_data.append(row_cells)
                row_cells = []

        # 補齊最後一行
        if row_cells:
            while len(row_cells) < COLS_PER_PAGE:
                row_cells.append("")
            table_data.append(row_cells)

        if not table_data:
            self._cleanup_temp_files()
            return output_path

        col_widths = [cell_w] * COLS_PER_PAGE
        row_heights = [cell_h] * len(table_data)

        table = Table(table_data, colWidths=col_widths, rowHeights=row_heights)
        table.setStyle(TableStyle([
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ]))

        doc.build([table])
        
        # 清理臨時文件
        self._cleanup_temp_files()
        
        return output_path

    def export_check_in_xlsx(
        self,
        households: List[Dict],
        filename: str = "報到.xlsx"
    ) -> str:
        """
        導出報到條碼到 Excel 文件
        
        格式：戶號 | 戶名 | 面積（坪） | 條碼
        
        Args:
            households: [
                {
                    'household_id': 'A106-02', 
                    'name': '洪正平', 
                    'share_amount': 129.03, 
                    'barcode_data': 'A106-02'
                }, 
                ...
            ]
            filename: 輸出文件名（默認為 報到.xlsx）
            
        Returns:
            輸出文件路徑
        """
        output_path = str(Path(self.output_dir) / filename)
        
        if OPENPYXL_AVAILABLE:
            # 使用 openpyxl
            from openpyxl import Workbook
            
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = '報到'
            
            # 設置列寬
            worksheet.column_dimensions['A'].width = 12
            worksheet.column_dimensions['B'].width = 18
            worksheet.column_dimensions['C'].width = 15
            worksheet.column_dimensions['D'].width = 18
            
            # 寫入標題
            headers = ['戶號', '戶名', '面積（坪）', '條碼']
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, size=11, color="FFFFFF")
            
            for col_idx, header in enumerate(headers, start=1):
                cell = worksheet.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            # 寫入數據
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row_idx, household in enumerate(households, start=2):
                # 戶號
                cell_a = worksheet.cell(row=row_idx, column=1, value=household['household_id'])
                cell_a.alignment = Alignment(horizontal='center', vertical='center')
                cell_a.border = thin_border
                
                # 戶名
                cell_b = worksheet.cell(row=row_idx, column=2, value=household['name'])
                cell_b.alignment = Alignment(horizontal='left', vertical='center')
                cell_b.border = thin_border
                
                # 面積（坪）
                share_amount = household.get('share_amount', 0.0)
                cell_c = worksheet.cell(row=row_idx, column=3, value=share_amount)
                cell_c.alignment = Alignment(horizontal='center', vertical='center')
                cell_c.number_format = '0.00'
                cell_c.border = thin_border
                
                # 條碼
                barcode_str = household.get('barcode_data', '')
                cell_d = worksheet.cell(row=row_idx, column=4, value=barcode_str)
                cell_d.alignment = Alignment(horizontal='center', vertical='center')
                cell_d.border = thin_border
            
            # 凍結標題行
            worksheet.freeze_panes = 'A2'
            
            workbook.save(output_path)
            
        elif PANDAS_AVAILABLE:
            # 使用 pandas
            import pandas as pd
            
            # 構建 DataFrame
            data = {
                '戶號': [h['household_id'] for h in households],
                '戶名': [h['name'] for h in households],
                '面積（坪）': [h.get('share_amount', 0.0) for h in households],
                '條碼': [h.get('barcode_data', '') for h in households]
            }
            df = pd.DataFrame(data)
            
            # 使用 ExcelWriter 設置樣式
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='報到', index=False)
                
                # 設置列寬
                worksheet = writer.sheets['報到']
                worksheet.column_dimensions['A'].width = 12
                worksheet.column_dimensions['B'].width = 18
                worksheet.column_dimensions['C'].width = 15
                worksheet.column_dimensions['D'].width = 18
        else:
            # 降級方案：使用 CSV
            import csv
            csv_path = output_path.replace('.xlsx', '.csv')
            
            with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.DictWriter(f, fieldnames=['戶號', '戶名', '面積（坪）', '條碼'])
                writer.writeheader()
                
                for household in households:
                    writer.writerow({
                        '戶號': household['household_id'],
                        '戶名': household['name'],
                        '面積（坪）': household.get('share_amount', 0.0),
                        '條碼': household.get('barcode_data', '')
                    })
            
            output_path = csv_path
        
        return output_path
